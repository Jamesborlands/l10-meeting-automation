#!/usr/bin/env python3
"""
Test just the conversion part on Render to see debug output
"""

import requests
import json

your_meeting_data = {
  "meeting_date": "2023-10-05",
  "attendees": ["Josh Keshen", "Lyndsey Dunnavant", "Josh Weinberg", "Regan Gentry", "Lisa Schick"],
  "new_commitments": [
    {
      "who": "Lyndsey Dunnavant",
      "task": "Ensure training policy is revised and prepared for compliance review.",
      "due_date": "Next meeting",
      "context": "To align with regulatory standards.",
      "dependencies": "Gather input from team and maintain policy visibility."
    },
    {
      "who": "Josh Weinberg",
      "task": "Monitor updates on audits and ensure compliance documentation is complete.",
      "due_date": "Next meeting",
      "context": "To facilitate smooth audits in future and ensure all states are compliant.",
      "dependencies": "Collaboration with state teams."
    }
  ],
  "issues_discussed": [
    {
      "issue": "The need to adjust the training policy.",
      "raised_by": "Lyndsey Dunnavant",
      "context": "The existing training policy is not being adhered to, and regulators may question training compliance.",
      "decision": "Lyndsey will redline the training policy for revision.",
      "owner": "Lyndsey Dunnavant"
    }
  ]
}

def test_render_conversion():
    """Test Render with minimal data to see debug output"""
    print("🧪 Testing Render conversion")
    print("="*50)
    
    payload = {
        "meeting_data": your_meeting_data,
        "excel_url": ""
    }
    
    print("Sending simplified data...")
    
    try:
        response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload,
            timeout=60
        )
        
        print(f"Response status: {response.status_code}")
        
        if response.status_code == 200:
            # Save and check the file
            with open('render_conversion_test.xlsx', 'wb') as f:
                f.write(response.content)
            
            print(f"File size: {len(response.content)} bytes")
            
            # Quick Excel check
            import openpyxl
            wb = openpyxl.load_workbook('render_conversion_test.xlsx')
            latest_sheet = wb[wb.sheetnames[-1]]
            
            print(f"New sheet: {latest_sheet.title}, Max row: {latest_sheet.max_row}")
            
            # Look for AI section specifically
            for row in range(45, min(65, latest_sheet.max_row + 1)):
                cell_value = latest_sheet.cell(row=row, column=1).value
                if cell_value and "AI IDENTIFIED" in str(cell_value):
                    print(f"AI section at row {row}")
                    
                    # Check the next few rows for content
                    for check_row in range(row + 1, min(row + 15, latest_sheet.max_row + 1)):
                        content = []
                        for col in range(1, 4):
                            val = latest_sheet.cell(row=check_row, column=col).value
                            if val:
                                content.append(str(val)[:30])
                        if content:
                            print(f"  Row {check_row}: {' | '.join(content)}")
                    break
            else:
                print("❌ No AI section found")
            
            wb.close()
            
        else:
            print(f"❌ Error: {response.status_code}")
            print(response.text)
    
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    test_render_conversion()
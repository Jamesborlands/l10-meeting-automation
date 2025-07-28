import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import re
from copy import copy

class L10SheetAutomation:
    """
    Automates L10 meeting workflow by duplicating sheets within the same workbook
    """
    
    def __init__(self, workbook_path):
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path)
        
    def get_latest_sheet(self):
        """Find the most recent L10 sheet in the workbook"""
        sheets = self.wb.sheetnames
        print(f"Found {len(sheets)} sheets: {sheets}")
        
        # Assume the last sheet is the most recent
        latest_sheet = self.wb[sheets[-1]]
        return latest_sheet
    
    def duplicate_sheet(self, source_sheet, new_date):
        """Duplicate a sheet and update the date"""
        # Create new sheet name (e.g., "Jul 23 2025" or match your format)
        new_sheet_name = new_date.strftime("%-m.%d.%Y")
        
        # Copy the sheet
        new_sheet = self.wb.copy_worksheet(source_sheet)
        new_sheet.title = new_sheet_name
        
        print(f"Created new sheet: {new_sheet_name}")
        
        # Update date in the new sheet (look for date patterns)
        for row in range(1, 10):  # Check first 10 rows
            for col in range(1, min(8, new_sheet.max_column + 1)):
                cell = new_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    # Look for date patterns and update them
                    if re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', str(cell.value)):
                        # Update to new date
                        cell.value = new_date.strftime("%m/%d/%Y")
                        print(f"Updated date in cell {row},{col}")
                        break
                    elif 'Day:' in str(cell.value):
                        # Update Day: field
                        cell.value = f"Day: {new_date.strftime('%m/%d/%Y')}"
                        print(f"Updated Day field in cell {row},{col}")
        
        return new_sheet
    
    def find_existing_todos(self, sheet):
        """Extract existing TO-DOs from the sheet"""
        existing_todos = []
        
        # Find TO-DO section
        todo_row = None
        for row in range(1, min(30, sheet.max_row)):
            for col in range(1, min(7, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and 'TO-DO' in str(cell_value).upper() and 'REVIEW' in str(cell_value).upper():
                    todo_row = row
                    break
            if todo_row:
                break
        
        if todo_row:
            # Look for TO-DO items after the header
            # Skip a few rows to get past headers
            for row in range(todo_row + 3, sheet.max_row + 1):
                who = sheet.cell(row=row, column=2).value  # WHO column
                todo = sheet.cell(row=row, column=3).value  # TO-DO column
                done = sheet.cell(row=row, column=4).value  # DONE? column
                notes = sheet.cell(row=row, column=5).value  # Notes column
                
                if who and todo:
                    existing_todos.append({
                        'WHO': str(who).strip(),
                        'TO-DO': str(todo).strip(),
                        'DONE?': str(done).strip() if done else '',
                        'NOTES': str(notes).strip() if notes else '',
                        'row': row
                    })
                elif not who and not todo and row > todo_row + 10:
                    break
        
        return existing_todos
    
    def add_ai_section(self, sheet, new_todos, new_issues, existing_todos=[]):
        """Add AI identified items section matching the exact format from screenshot"""
        # Validate and sanitize inputs
        if not isinstance(new_todos, list):
            print(f"WARNING: new_todos is not a list, got {type(new_todos)}")
            new_todos = []
        
        if not isinstance(new_issues, list):
            print(f"WARNING: new_issues is not a list, got {type(new_issues)}")
            new_issues = []
            
        if not isinstance(existing_todos, list):
            existing_todos = []
        
        print(f"Adding AI section with {len(new_todos)} TODOs, {len(new_issues)} issues, and {len(existing_todos)} existing TODOs")
        
        # Find the last row with content
        last_row = sheet.max_row
        
        # Add some space
        start_row = last_row + 3
        
        # Add main header with blue background and white text
        header_cell = sheet.cell(row=start_row, column=1, 
                                value="AI IDENTIFIED ITEMS (Review & Move to Appropriate Sections)")
        header_cell.font = Font(bold=True, color="FFFFFF", size=12)
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Merge cells for main header across all columns
        sheet.merge_cells(start_row=start_row, start_column=1, 
                         end_row=start_row, end_column=5)
        
        current_row = start_row + 1
        
        # Add TO-DO section with blue headers
        if new_todos or existing_todos:
            # TO-DO section headers with blue background
            headers = ['Who', "To-do's", 'When', 'Context', 'Dependencies']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Add new TODOs
            for todo in new_todos:
                try:
                    if not isinstance(todo, dict):
                        continue
                    
                    sheet.cell(row=current_row, column=1, value=str(todo.get('WHO', '')))
                    sheet.cell(row=current_row, column=2, value=str(todo.get('TO-DO', '')))
                    sheet.cell(row=current_row, column=3, value=str(todo.get('DUE DATE', todo.get('WHEN', 'Next meeting'))))
                    sheet.cell(row=current_row, column=4, value=str(todo.get('CONTEXT', '')))
                    sheet.cell(row=current_row, column=5, value=str(todo.get('DEPENDENCIES', 'None')))
                    
                    current_row += 1
                except Exception as e:
                    print(f"Error processing TODO: {e}")
                    continue
        
        # Add space before Issue List
        current_row += 1
        
        # Add Issue List section with blue headers
        if new_issues:
            # Issue List header
            issue_header = sheet.cell(row=current_row, column=1, value="Issue List")
            issue_header.font = Font(bold=True, color="FFFFFF")
            issue_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            issue_header.alignment = Alignment(horizontal="left", vertical="center")
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            current_row += 1
            
            # Issue section headers with blue background
            issue_headers = ['Issue_description', 'Raised By', 'Issue Cause', 'Related Discussions', 'Notes']
            for col, header in enumerate(issue_headers, 1):
                cell = sheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Add issues
            for issue in new_issues:
                try:
                    if not isinstance(issue, dict):
                        continue
                    
                    sheet.cell(row=current_row, column=1, value=str(issue.get('issue_description', issue.get('ISSUE', ''))))
                    sheet.cell(row=current_row, column=2, value=str(issue.get('who_raised_it', issue.get('RAISED BY', ''))))
                    sheet.cell(row=current_row, column=3, value=str(issue.get('root_cause', issue.get('ISSUE CAUSE', ''))))
                    sheet.cell(row=current_row, column=4, value=str(issue.get('related_discussions', issue.get('RELATED DISCUSSIONS', ''))))
                    sheet.cell(row=current_row, column=5, value=str(issue.get('notes', issue.get('NOTES', ''))))
                    
                    current_row += 1
                except Exception as e:
                    print(f"Error processing Issue: {e}")
                    continue
        
        # Add space before Todo Review
        current_row += 1
        
        # Add Todo Review section with blue headers
        if existing_todos:
            # Todo Review header
            review_header = sheet.cell(row=current_row, column=1, value="Todo Review")
            review_header.font = Font(bold=True, color="FFFFFF")
            review_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            review_header.alignment = Alignment(horizontal="left", vertical="center")
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1
            
            # Todo Review section headers
            review_headers = ['Who', 'Todo', 'Status', 'Notes']
            for col, header in enumerate(review_headers, 1):
                cell = sheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Add existing todos for review
            for todo in existing_todos:
                try:
                    if not isinstance(todo, dict):
                        continue
                    
                    sheet.cell(row=current_row, column=1, value=str(todo.get('WHO', '')))
                    sheet.cell(row=current_row, column=2, value=str(todo.get('TO-DO', '')))
                    sheet.cell(row=current_row, column=3, value=str(todo.get('DONE?', 'In Progress')))
                    sheet.cell(row=current_row, column=4, value=str(todo.get('NOTES', '')))
                    
                    current_row += 1
                except Exception as e:
                    print(f"Error processing existing TODO: {e}")
                    continue
        
        print(f"Added AI section with {current_row - start_row} total rows")
        return current_row
    
    def update_current_sheet_with_ai_data(self, meeting_data):
        """Update the current sheet with AI-identified items instead of creating new sheet"""
        print("=== L10 SHEET AUTOMATION (Update Current Sheet) ===")
        
        # Get the latest sheet to update
        current_sheet = self.get_latest_sheet()
        print(f"Updating sheet: {current_sheet.title}")
        
        # Find existing TO-DOs in the current sheet
        existing_todos = self.find_existing_todos(current_sheet)
        print(f"Found {len(existing_todos)} existing TO-DOs")
        
        # Get new TO-DOs from meeting data directly
        new_todos_from_meeting = meeting_data.get('NEW TO-DOS', [])
        
        # FAILSAFE: Also check for alternative formats
        if not new_todos_from_meeting and 'new_commitments' in meeting_data:
            print("FAILSAFE: Converting new_commitments to NEW TO-DOS")
            new_todos_from_meeting = []
            for commitment in meeting_data['new_commitments']:
                new_todos_from_meeting.append({
                    'WHO': commitment.get('who', ''),
                    'TO-DO': commitment.get('task', ''),
                    'WHEN': commitment.get('due_date', 'Next meeting'),
                    'CONTEXT': commitment.get('context', ''),
                    'DEPENDENCIES': commitment.get('dependencies', 'None')
                })
        
        # Filter out duplicates
        truly_new_todos = []
        for new_todo in new_todos_from_meeting:
            is_duplicate = False
            for existing in existing_todos:
                if (new_todo.get('WHO', '').lower() == existing.get('WHO', '').lower() and
                    new_todo.get('TO-DO', '').lower() in existing.get('TO-DO', '').lower()):
                    is_duplicate = True
                    break
            if not is_duplicate:
                truly_new_todos.append(new_todo)
        
        print(f"Found {len(truly_new_todos)} truly new TO-DOs")
        
        # Add AI section with new items
        new_issues = meeting_data.get('ISSUES LIST (IDS)', [])
        
        # FAILSAFE: Also check for alternative formats
        if not new_issues and 'issues_discussed' in meeting_data:
            print("FAILSAFE: Converting issues_discussed to ISSUES LIST (IDS)")
            new_issues = []
            for issue in meeting_data['issues_discussed']:
                new_issues.append({
                    'issue_description': issue.get('issue', ''),
                    'who_raised_it': issue.get('raised_by', ''),
                    'root_cause': issue.get('context', ''),
                    'related_discussions': ', '.join(issue.get('discussion_points', [])) if issue.get('discussion_points') else '',
                    'notes': f"Decision: {issue.get('decision', '')} | Owner: {issue.get('owner', '')}"
                })
        
        # ULTIMATE FAILSAFE: If still no data, create debug entry
        if not truly_new_todos and not new_issues:
            print("WARNING: No data found! Adding debug entry")
            truly_new_todos = [{
                'WHO': 'System',
                'TO-DO': 'DEBUG: No meeting data was found - check Zapier payload structure',
                'WHEN': 'Immediate',
                'CONTEXT': f'Meeting data keys: {list(meeting_data.keys())}',
                'DEPENDENCIES': 'Check /echo endpoint with same payload'
            }]
        
        # Add AI section with proper formatting and include existing todos for review
        self.add_ai_section(current_sheet, truly_new_todos, new_issues, existing_todos)
        
        # Save the workbook
        self.wb.save(self.workbook_path)
        print(f"Updated sheet: {current_sheet.title} with AI section")
        
        return {
            'sheet_name': current_sheet.title,
            'new_todos_count': len(truly_new_todos),
            'new_issues_count': len(new_issues),
            'existing_todos_count': len(existing_todos)
        }
    
    def process_meeting_output(self, meeting_text):
        """Parse the meeting output text"""
        from l10_processor import parse_l10_text
        return parse_l10_text(meeting_text)
    
    def create_next_l10_sheet(self, meeting_output_file, meeting_cadence='weekly'):
        """
        Main automation function:
        1. Duplicate the latest sheet
        2. Update the date
        3. Keep existing TO-DOs (including completed ones)
        4. Add AI identified items section
        5. Save the workbook
        """
        print("=== L10 SHEET AUTOMATION ===")
        
        # Get the latest sheet
        latest_sheet = self.get_latest_sheet()
        print(f"Using sheet: {latest_sheet.title}")
        
        # Calculate next meeting date
        today = datetime.now()
        if meeting_cadence == 'weekly':
            next_date = today + timedelta(days=7)
        else:
            next_date = today + timedelta(days=14)
        
        # Duplicate the sheet
        new_sheet = self.duplicate_sheet(latest_sheet, next_date)
        
        # Parse the meeting output
        with open(meeting_output_file, 'r') as f:
            meeting_text = f.read()
        
        meeting_data = self.process_meeting_output(meeting_text)
        
        # Find existing TO-DOs in the new sheet
        existing_todos = self.find_existing_todos(new_sheet)
        print(f"Found {len(existing_todos)} existing TO-DOs")
        
        # Get new TO-DOs from meeting
        new_todos_from_meeting = meeting_data.get('NEW TO-DOS', [])
        
        # Filter out duplicates
        truly_new_todos = []
        for new_todo in new_todos_from_meeting:
            is_duplicate = False
            for existing in existing_todos:
                if (new_todo.get('WHO', '').lower() == existing.get('WHO', '').lower() and
                    new_todo.get('TO-DO', '').lower() in existing.get('TO-DO', '').lower()):
                    is_duplicate = True
                    break
            if not is_duplicate:
                truly_new_todos.append(new_todo)
        
        print(f"Found {len(truly_new_todos)} truly new TO-DOs")
        
        # Add AI section with new items
        new_issues = meeting_data.get('ISSUES LIST (IDS)', [])
        
        self.add_ai_section(new_sheet, truly_new_todos, new_issues)
        
        # Save the workbook
        self.wb.save(self.workbook_path)
        print(f"Saved workbook with new sheet: {new_sheet.title}")
        
        return {
            'new_sheet_name': new_sheet.title,
            'next_date': next_date.strftime("%m/%d/%Y"),
            'new_todos_count': len(truly_new_todos),
            'new_issues_count': len(new_issues),
            'existing_todos_count': len(existing_todos)
        }
    
    def create_next_l10_sheet_from_data(self, meeting_data, meeting_cadence='weekly'):
        """Process meeting data directly without text file"""
        print("=== L10 SHEET AUTOMATION (Direct Data) ===")
        
        # Get the latest sheet
        latest_sheet = self.get_latest_sheet()
        print(f"Using sheet: {latest_sheet.title}")
        
        # Calculate next meeting date
        today = datetime.now()
        if meeting_cadence == 'weekly':
            next_date = today + timedelta(days=7)
        else:
            next_date = today + timedelta(days=14)
        
        # Duplicate the sheet
        new_sheet = self.duplicate_sheet(latest_sheet, next_date)
        
        # Find existing TO-DOs in the new sheet
        existing_todos = self.find_existing_todos(new_sheet)
        print(f"Found {len(existing_todos)} existing TO-DOs")
        
        # Get new TO-DOs from meeting data directly
        new_todos_from_meeting = meeting_data.get('NEW TO-DOS', [])
        
        # FAILSAFE: Also check for alternative formats
        if not new_todos_from_meeting and 'new_commitments' in meeting_data:
            print("FAILSAFE: Converting new_commitments to NEW TO-DOS")
            new_todos_from_meeting = []
            for commitment in meeting_data['new_commitments']:
                new_todos_from_meeting.append({
                    'WHO': commitment.get('who', ''),
                    'TO-DO': commitment.get('task', ''),
                    'DUE DATE': commitment.get('due_date', ''),
                    'CONTEXT': commitment.get('context', ''),
                    'DEPENDENCIES': commitment.get('dependencies', '')
                })
        
        # Filter out duplicates
        truly_new_todos = []
        for new_todo in new_todos_from_meeting:
            is_duplicate = False
            for existing in existing_todos:
                if (new_todo.get('WHO', '').lower() == existing.get('WHO', '').lower() and
                    new_todo.get('TO-DO', '').lower() in existing.get('TO-DO', '').lower()):
                    is_duplicate = True
                    break
            if not is_duplicate:
                truly_new_todos.append(new_todo)
        
        print(f"Found {len(truly_new_todos)} truly new TO-DOs")
        
        # Add AI section with new items
        new_issues = meeting_data.get('ISSUES LIST (IDS)', [])
        
        # FAILSAFE: Also check for alternative formats
        if not new_issues and 'issues_discussed' in meeting_data:
            print("FAILSAFE: Converting issues_discussed to ISSUES LIST (IDS)")
            new_issues = []
            for issue in meeting_data['issues_discussed']:
                new_issues.append({
                    'issue_description': issue.get('issue', ''),
                    'who_raised_it': issue.get('raised_by', ''),
                    'root_cause': issue.get('context', ''),
                    'related_discussions': ', '.join(issue.get('discussion_points', [])) if issue.get('discussion_points') else '',
                    'notes': f"Decision: {issue.get('decision', '')} | Owner: {issue.get('owner', '')}"
                })
        
        # ULTIMATE FAILSAFE: If still no data, create debug entry
        if not truly_new_todos and not new_issues:
            print("WARNING: No data found! Adding debug entry")
            truly_new_todos = [{
                'WHO': 'System',
                'TO-DO': 'DEBUG: No meeting data was found - check Zapier payload structure',
                'DUE DATE': 'Immediate',
                'CONTEXT': f'Meeting data keys: {list(meeting_data.keys())}',
                'DEPENDENCIES': 'Check /echo endpoint with same payload'
            }]
        
        self.add_ai_section(new_sheet, truly_new_todos, new_issues, existing_todos)
        
        # Save the workbook
        self.wb.save(self.workbook_path)
        print(f"Saved workbook with new sheet: {new_sheet.title}")
        
        return {
            'new_sheet_name': new_sheet.title,
            'next_date': next_date.strftime("%m/%d/%Y"),
            'new_todos_count': len(truly_new_todos),
            'new_issues_count': len(new_issues),
            'existing_todos_count': len(existing_todos)
        }
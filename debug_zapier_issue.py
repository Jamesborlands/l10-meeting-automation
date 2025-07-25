#!/usr/bin/env python3
"""
Comprehensive debugging script to find why Zapier AI section is empty
"""

import json
import requests
import openpyxl
from datetime import datetime

def test_exact_zapier_structure():
    """Test different possible Zapier data structures"""
    print("🔍 TESTING DIFFERENT ZAPIER DATA STRUCTURES")
    print("="*60)
    
    # Test different possible wrapper structures
    test_cases = [
        {
            "name": "Direct Data (as shown in your example)",
            "payload": {
                "meeting_data": {
                    "meeting_date": "2023-10-24",
                    "attendees": ["Josh Keshen", "Lyndsey Dunnavant"],
                    "NEW TO-DOS": [
                        {
                            "WHO": "Test User",
                            "TO-DO": "Test task for debugging",
                            "DUE DATE": "Next meeting",
                            "CONTEXT": "Testing",
                            "DEPENDENCIES": "None"
                        }
                    ],
                    "ISSUES LIST (IDS)": [
                        {
                            "issue_description": "Test issue",
                            "who_raised_it": "Test User",
                            "root_cause": "Testing",
                            "related_discussions": "Test discussion",
                            "notes": "Test notes"
                        }
                    ]
                }
            }
        },
        {
            "name": "Wrapped in array",
            "payload": {
                "meeting_data": [{
                    "NEW TO-DOS": [
                        {
                            "WHO": "Test User",
                            "TO-DO": "Test task for debugging",
                            "DUE DATE": "Next meeting",
                            "CONTEXT": "Testing",
                            "DEPENDENCIES": "None"
                        }
                    ]
                }]
            }
        },
        {
            "name": "String wrapped",
            "payload": {
                "meeting_data": json.dumps({
                    "NEW TO-DOS": [
                        {
                            "WHO": "Test User",
                            "TO-DO": "Test task for debugging",
                            "DUE DATE": "Next meeting",
                            "CONTEXT": "Testing",
                            "DEPENDENCIES": "None"
                        }
                    ]
                })
            }
        },
        {
            "name": "Nested data structure",
            "payload": {
                "data": {
                    "meeting_data": {
                        "NEW TO-DOS": [
                            {
                                "WHO": "Test User",
                                "TO-DO": "Test task for debugging",
                                "DUE DATE": "Next meeting",
                                "CONTEXT": "Testing",
                                "DEPENDENCIES": "None"
                            }
                        ]
                    }
                }
            }
        }
    ]
    
    for test_case in test_cases:
        print(f"\nTesting: {test_case['name']}")
        print("-" * 40)
        
        try:
            response = requests.post(
                'https://l10-meeting-automation-29fl.onrender.com/process-l10',
                json=test_case['payload'],
                timeout=30
            )
            
            print(f"Response status: {response.status_code}")
            
            if response.status_code == 200:
                # Save and analyze
                filename = f"debug_{test_case['name'].replace(' ', '_')}.xlsx"
                with open(filename, 'wb') as f:
                    f.write(response.content)
                
                # Quick check
                wb = openpyxl.load_workbook(filename)
                sheet = wb[wb.sheetnames[-1]]
                
                ai_found = False
                items_found = 0
                
                for row in range(45, min(70, sheet.max_row + 1)):
                    cell = sheet.cell(row=row, column=1).value
                    if cell and "AI IDENTIFIED" in str(cell):
                        ai_found = True
                        # Count items in next 10 rows
                        for r in range(row + 1, min(row + 10, sheet.max_row + 1)):
                            if sheet.cell(row=r, column=2).value:
                                val = str(sheet.cell(row=r, column=2).value)
                                if "Test" in val or "task" in val:
                                    items_found += 1
                        break
                
                wb.close()
                
                if ai_found and items_found > 0:
                    print(f"✅ SUCCESS: AI section found with {items_found} items")
                elif ai_found:
                    print(f"⚠️  AI section found but EMPTY")
                else:
                    print(f"❌ No AI section found")
                    
            else:
                print(f"❌ Error: {response.status_code}")
                print(response.text[:200])
                
        except Exception as e:
            print(f"❌ Exception: {e}")

def capture_raw_zapier_webhook():
    """Instructions for capturing raw Zapier data"""
    print("\n" + "="*60)
    print("📋 TO CAPTURE RAW ZAPIER DATA:")
    print("="*60)
    print("""
1. Use a webhook testing service:
   - Go to https://webhook.site
   - Copy the unique URL
   - Temporarily point your Zapier webhook to this URL
   - Run your Zapier flow
   - Copy the EXACT raw payload

2. Or add logging to app.py:
   Add this at the very start of process_l10():
   
   with open(f'/tmp/zapier_raw_{datetime.now().timestamp()}.json', 'w') as f:
       json.dump(request.get_json(force=True), f, indent=2)
   
3. Share the exact structure you see!
""")

if __name__ == "__main__":
    test_exact_zapier_structure()
    capture_raw_zapier_webhook()
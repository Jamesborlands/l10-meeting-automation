#!/usr/bin/env python3
"""
Debug the Render deployment to see if it has the latest conversion code
"""

import requests
import json

def test_render_debug_endpoint():
    """Check what's on Render"""
    print("🔍 DEBUGGING RENDER DEPLOYMENT")
    print("="*50)
    
    try:
        # Check debug endpoint
        response = requests.get('https://l10-meeting-automation-29fl.onrender.com/debug', timeout=10)
        if response.status_code == 200:
            debug_info = response.json()
            print(f"✅ Render current directory: {debug_info.get('current_dir')}")
            print(f"✅ Files on Render: {debug_info.get('files')}")
            print(f"✅ Excel files: {debug_info.get('xlsx_files')}")
        else:
            print(f"❌ Debug endpoint failed: {response.status_code}")
            
    except Exception as e:
        print(f"❌ Debug check error: {e}")

def test_simple_conversion():
    """Test with very simple data to isolate the issue"""
    print("\n🧪 TESTING SIMPLE CONVERSION")
    print("="*50)
    
    # Minimal test data
    simple_data = {
        "new_commitments": [
            {
                "who": "Test User",
                "task": "Test task for debugging",
                "due_date": "Next week",
                "context": "Testing conversion",
                "dependencies": "None"
            }
        ],
        "issues_discussed": [
            {
                "issue": "Test issue for debugging",
                "raised_by": "Test User",
                "context": "Testing issue conversion",
                "decision": "Test decision",
                "owner": "Test User"
            }
        ]
    }
    
    payload = {
        "meeting_data": simple_data,
        "excel_url": ""
    }
    
    print("Sending minimal test data:")
    print(f"  - 1 new commitment")
    print(f"  - 1 issue discussed")
    
    try:
        response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload,
            timeout=60
        )
        
        print(f"\nResponse status: {response.status_code}")
        
        if response.status_code == 200:
            # Save and analyze
            with open('simple_render_test.xlsx', 'wb') as f:
                f.write(response.content)
            
            print(f"✅ File saved: simple_render_test.xlsx ({len(response.content)} bytes)")
            
            # Quick analysis
            import openpyxl
            wb = openpyxl.load_workbook('simple_render_test.xlsx')
            latest_sheet = wb[wb.sheetnames[-1]]
            
            print(f"✅ Latest sheet: {latest_sheet.title}")
            print(f"✅ Max row: {latest_sheet.max_row}")
            
            # Look for AI section and content
            ai_found = False
            items_found = 0
            
            for row in range(1, latest_sheet.max_row + 1):
                cell_value = latest_sheet.cell(row=row, column=1).value
                if cell_value and "AI IDENTIFIED" in str(cell_value):
                    ai_found = True
                    print(f"✅ AI section found at row {row}")
                    
                    # Check next 10 rows for content
                    for check_row in range(row + 1, min(row + 11, latest_sheet.max_row + 1)):
                        row_content = []
                        for col in range(1, 6):
                            val = latest_sheet.cell(row=check_row, column=col).value
                            if val and str(val).strip():
                                row_content.append(f"C{col}:{str(val)[:20]}...")
                        
                        if row_content:
                            items_found += 1
                            print(f"  Row {check_row}: {' | '.join(row_content)}")
                    break
            
            if ai_found:
                if items_found > 0:
                    print(f"🎉 SUCCESS: Found {items_found} items in AI section")
                else:
                    print("⚠️  AI section found but NO ITEMS populated")
            else:
                print("❌ AI section not found at all")
            
            wb.close()
            
        else:
            print(f"❌ Request failed: {response.status_code}")
            print(f"Error: {response.text}")
            
    except Exception as e:
        print(f"❌ Request error: {e}")

def test_original_format():
    """Test with original format to see if that works"""
    print("\n🧪 TESTING ORIGINAL L10 FORMAT")
    print("="*50)
    
    # Use the original working format
    original_data = {
        "NEW TO-DOS": [
            {
                "WHO": "Test User",
                "TO-DO": "Test original format task",
                "DUE DATE": "Next week",
                "CONTEXT": "Testing original format",
                "DEPENDENCIES": "None"
            }
        ],
        "ISSUES LIST (IDS)": [
            {
                "issue_description": "Test original format issue",
                "who_raised_it": "Test User",
                "root_cause": "Testing",
                "related_discussions": "Testing discussion",
                "notes": "Test notes"
            }
        ]
    }
    
    payload = {
        "meeting_data": original_data,
        "excel_url": ""
    }
    
    try:
        response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload,
            timeout=60
        )
        
        print(f"Response status: {response.status_code}")
        
        if response.status_code == 200:
            with open('original_format_test.xlsx', 'wb') as f:
                f.write(response.content)
            
            print(f"✅ Original format file saved ({len(response.content)} bytes)")
            
            # Quick check
            import openpyxl
            wb = openpyxl.load_workbook('original_format_test.xlsx')
            latest_sheet = wb[wb.sheetnames[-1]]
            
            # Count items in AI section
            items_count = 0
            for row in range(50, min(70, latest_sheet.max_row + 1)):
                for col in range(1, 6):
                    val = latest_sheet.cell(row=row, column=col).value
                    if val and str(val).strip():
                        if "Test" in str(val):
                            items_count += 1
                            break
            
            print(f"✅ Original format items found: {items_count}")
            wb.close()
            
        else:
            print(f"❌ Original format failed: {response.status_code}")
            
    except Exception as e:
        print(f"❌ Original format error: {e}")

if __name__ == "__main__":
    test_render_debug_endpoint()
    test_simple_conversion()
    test_original_format()
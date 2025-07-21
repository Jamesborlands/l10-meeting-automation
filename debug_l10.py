import openpyxl
from datetime import datetime
import os

def debug_template_and_process():
    """Debug version that shows what's happening at each step"""
    
    # First, let's see what's in the template
    print("=== ANALYZING TEMPLATE ===")
    wb = openpyxl.load_workbook('L10 Summary Template 1.xlsx')
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns\n")
    
    # Map out where sections are
    sections_found = {}
    for row in range(1, min(40, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # Look for section headers
                if any(keyword in cell_str.upper() for keyword in ['HEADLINE', 'TO-DO', 'REVIEW', 'ISSUES', 'IDS', 'RATE', 'RATING']):
                    print(f"Row {row}, Col {col}: {cell_value}")
                    sections_found[cell_str] = (row, col)
    
    print("\n=== PROCESSING DATA ===")
    
    # Read the L10 output
    with open('l10_output.txt', 'r') as f:
        text_input = f.read()
    
    # Parse sections
    sections = parse_l10_text_debug(text_input)
    
    # Show what we parsed
    for section, content in sections.items():
        if content and content != 'No metrics discussed.' and content != 'No quarterly goals discussed.' and content != 'None discussed.':
            print(f"\n{section}:")
            if isinstance(content, list):
                print(f"  Found {len(content)} items")
                if len(content) > 0 and isinstance(content[0], dict):
                    print(f"  First item: {content[0]}")
            else:
                print(f"  Content: {content}")
    
    return sections_found, sections

def parse_l10_text_debug(text):
    """Debug version of parse function with print statements"""
    sections = {
        'HEADLINES': [],
        'TO-DO REVIEW': [],
        'SCORECARD/METRICS REVIEW': '',
        'ROCK REVIEW': '',
        'CUSTOMER/EMPLOYEE HEADLINES': [],
        'ISSUES LIST (IDS)': [],
        'NEW TO-DOS': [],
        'CASCADING MESSAGES': '',
        'MEETING RATING': []
    }
    
    current_section = None
    current_item = {}
    lines = text.strip().split('\n')
    
    print(f"Total lines to process: {len(lines)}")
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        if not line:
            continue
            
        # Check for section headers
        if line.startswith('**') and line.endswith('**'):
            if current_item and current_section in ['TO-DO REVIEW', 'ISSUES LIST (IDS)', 'NEW TO-DOS']:
                sections[current_section].append(current_item)
                current_item = {}
            
            current_section = line.strip('*').strip()
            print(f"Found section: {current_section}")
            continue
        
        # Handle different sections
        if current_section == 'TO-DO REVIEW':
            if line == '---':
                if current_item:
                    sections['TO-DO REVIEW'].append(current_item)
                    print(f"  Added TO-DO: {current_item}")
                    current_item = {}
            elif line.startswith('WHO:'):
                current_item['WHO'] = line[4:].strip()
            elif line.startswith('TO-DO:'):
                current_item['TO-DO'] = line[6:].strip()
            elif line.startswith('DONE?:'):
                current_item['DONE?'] = line[6:].strip()
            elif line.startswith('NOTES:'):
                current_item['NOTES'] = line[6:].strip()
    
    # Save any final pending item
    if current_item and current_section in ['TO-DO REVIEW', 'ISSUES LIST (IDS)', 'NEW TO-DOS']:
        sections[current_section].append(current_item)
    
    return sections

# Run the debug
if __name__ == "__main__":
    template_sections, parsed_data = debug_template_and_process()
#!/usr/bin/env python3
"""
Quick script to verify the Excel output has the AI section populated
"""

import openpyxl

def verify_excel_file(filename):
    """Verify the Excel file has the AI section with items"""
    print(f"🔍 Verifying Excel file: {filename}")
    
    try:
        wb = openpyxl.load_workbook(filename)
        sheets = wb.sheetnames
        print(f"✅ Workbook loaded with {len(sheets)} sheets")
        
        # Get the latest sheet (last one)
        latest_sheet = wb[sheets[-1]]
        sheet_name = latest_sheet.title
        print(f"✅ Latest sheet: {sheet_name}")
        print(f"✅ Sheet dimensions: {latest_sheet.max_row} rows x {latest_sheet.max_column} columns")
        
        # Look for AI section
        ai_section_found = False
        ai_items_found = 0
        
        for row in range(1, latest_sheet.max_row + 1):
            for col in range(1, latest_sheet.max_column + 1):
                cell_value = latest_sheet.cell(row=row, column=col).value
                if cell_value and "AI IDENTIFIED ITEMS" in str(cell_value):
                    ai_section_found = True
                    ai_section_row = row
                    print(f"✅ Found AI section at row {row}")
                    
                    # Count items in the AI section (look for the next 20 rows)
                    for check_row in range(row + 1, min(row + 21, latest_sheet.max_row + 1)):
                        check_value = latest_sheet.cell(row=check_row, column=2).value
                        if check_value and str(check_value).strip():
                            ai_items_found += 1
                    
                    break
            if ai_section_found:
                break
        
        if ai_section_found:
            print(f"✅ AI section found with approximately {ai_items_found} items")
            if ai_items_found > 0:
                print("🎉 SUCCESS: AI section is populated with items!")
            else:
                print("⚠️  AI section found but appears empty")
        else:
            print("❌ AI section not found")
        
        wb.close()
        return ai_section_found and ai_items_found > 0
        
    except Exception as e:
        print(f"❌ Error verifying Excel file: {e}")
        return False

if __name__ == "__main__":
    # Check both local and render outputs
    files_to_check = ['test_local_output.xlsx', 'test_render_output.xlsx']
    
    for filename in files_to_check:
        if filename == 'test_local_output.xlsx':
            print("\n" + "="*50)
            print("LOCAL TEST OUTPUT")
            print("="*50)
        else:
            print("\n" + "="*50)
            print("RENDER TEST OUTPUT")
            print("="*50)
        
        verify_excel_file(filename)
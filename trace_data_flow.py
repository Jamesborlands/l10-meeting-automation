#!/usr/bin/env python3
"""
Test the exact same flow locally to trace where the data gets lost
"""

import json
import shutil
from l10_sheet_automation import L10SheetAutomation
from l10_processor import parse_l10_json

your_data = {
    "new_commitments": [
        {
            "who": "Lyndsey Dunnavant",
            "task": "Ensure training policy is revised and prepared for compliance review.",
            "due_date": "Next meeting",
            "context": "To align with regulatory standards.",
            "dependencies": "Gather input from team and maintain policy visibility."
        }
    ],
    "issues_discussed": [
        {
            "issue": "The need to adjust the training policy.",
            "raised_by": "Lyndsey Dunnavant",
            "context": "The existing training policy is not being adhered to",
            "decision": "Lyndsey will redline the training policy for revision.",
            "owner": "Lyndsey Dunnavant"
        }
    ]
}

def trace_local_flow():
    """Trace the exact same flow locally"""
    print("🔍 TRACING LOCAL DATA FLOW")
    print("="*50)
    
    # Step 1: Parse the data (same as app.py)
    print("Step 1: Parse meeting data")
    print(f"Raw data keys: {list(your_data.keys())}")
    
    meeting_data = parse_l10_json(your_data)
    print(f"After parse_l10_json keys: {list(meeting_data.keys())}")
    print(f"NEW TO-DOS count: {len(meeting_data.get('NEW TO-DOS', []))}")
    print(f"ISSUES count: {len(meeting_data.get('ISSUES LIST (IDS)', []))}")
    
    if meeting_data.get('NEW TO-DOS'):
        print("NEW TO-DOS content:")
        for i, todo in enumerate(meeting_data['NEW TO-DOS']):
            print(f"  {i+1}. {todo}")
    
    if meeting_data.get('ISSUES LIST (IDS)'):
        print("ISSUES content:")
        for i, issue in enumerate(meeting_data['ISSUES LIST (IDS)']):
            print(f"  {i+1}. {issue}")
    
    # Step 2: Create automation (same as app.py)
    print(f"\nStep 2: Create L10SheetAutomation")
    shutil.copy('L10 Summary Template 1.xlsx', 'trace_test.xlsx')
    automation = L10SheetAutomation('trace_test.xlsx')
    
    # Step 3: Run automation (same as app.py)
    print(f"\nStep 3: Run create_next_l10_sheet_from_data")
    result = automation.create_next_l10_sheet_from_data(meeting_data, 'weekly')
    
    print(f"\nAutomation result: {result}")
    
    # Step 4: Verify the Excel output
    print(f"\nStep 4: Verify Excel output")
    import openpyxl
    wb = openpyxl.load_workbook('trace_test.xlsx')
    latest_sheet = wb[wb.sheetnames[-1]]
    
    print(f"Latest sheet: {latest_sheet.title}")
    print(f"Max row: {latest_sheet.max_row}")
    
    # Look for AI section and items
    for row in range(50, min(70, latest_sheet.max_row + 1)):
        cell_value = latest_sheet.cell(row=row, column=1).value
        if cell_value and "AI IDENTIFIED" in str(cell_value):
            print(f"\nAI section found at row {row}")
            
            # Show the next several rows
            for check_row in range(row, min(row + 15, latest_sheet.max_row + 1)):
                row_content = []
                for col in range(1, 6):
                    val = latest_sheet.cell(row=check_row, column=col).value
                    if val:
                        row_content.append(f"C{col}:{str(val)[:25]}...")
                
                if row_content:
                    print(f"  Row {check_row}: {' | '.join(row_content)}")
            break
    
    wb.close()

if __name__ == "__main__":
    trace_local_flow()
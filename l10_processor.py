import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
from copy import copy
import re

def parse_l10_json(input_data):
    """Parse L10 meeting data - handles JSON input"""
    
    # If it's a string, parse it as JSON
    if isinstance(input_data, str):
        try:
            # Clean up if wrapped in backticks
            if '```' in input_data:
                input_data = input_data.split('```')[1]
                if input_data.startswith('json\n'):
                    input_data = input_data[5:]
            
            # Parse JSON
            parsed_data = json.loads(input_data.strip())
            return convert_to_l10_format(parsed_data)
        except json.JSONDecodeError as e:
            print(f"JSON parse error: {e}")
            # Fall back to text parsing
            return parse_l10_text(input_data)
    
    # If it's already a dict, convert it to L10 format
    return convert_to_l10_format(input_data)

def convert_to_l10_format(data):
    """Convert various JSON formats to L10 format"""
    print(f"=== DEBUG: Converting data format ===")
    print(f"Input keys: {list(data.keys()) if isinstance(data, dict) else 'Not a dict'}")
    
    # If it's already in L10 format, return as-is
    if 'NEW TO-DOS' in data or 'ISSUES LIST (IDS)' in data:
        print("Already in L10 format")
        return data
    
    # Convert from alternative format
    converted = {}
    
    # Map new_commitments to NEW TO-DOS
    if 'new_commitments' in data:
        converted['NEW TO-DOS'] = []
        for commitment in data['new_commitments']:
            todo_item = {
                'WHO': commitment.get('who', ''),
                'TO-DO': commitment.get('task', ''),
                'DUE DATE': commitment.get('due_date', ''),
                'CONTEXT': commitment.get('context', ''),
                'DEPENDENCIES': commitment.get('dependencies', '')
            }
            converted['NEW TO-DOS'].append(todo_item)
        print(f"Converted {len(converted['NEW TO-DOS'])} new commitments to NEW TO-DOS")
    
    # Map issues_discussed to ISSUES LIST (IDS)
    if 'issues_discussed' in data:
        converted['ISSUES LIST (IDS)'] = []
        for issue in data['issues_discussed']:
            issue_item = {
                'issue_description': issue.get('issue', ''),
                'who_raised_it': issue.get('raised_by', ''),
                'root_cause': issue.get('context', ''),
                'related_discussions': ', '.join(issue.get('discussion_points', [])) if issue.get('discussion_points') else '',
                'notes': f"Decision: {issue.get('decision', '')} | Owner: {issue.get('owner', '')}"
            }
            converted['ISSUES LIST (IDS)'].append(issue_item)
        print(f"Converted {len(converted['ISSUES LIST (IDS)'])} issues to ISSUES LIST (IDS)")
    
    # Map todo_review to TO-DO REVIEW
    if 'todo_review' in data:
        converted['TO-DO REVIEW'] = []
        for todo in data['todo_review']:
            todo_item = {
                'WHO': todo.get('who', ''),
                'TO-DO': todo.get('todo', ''),
                'DONE?': 'Yes' if todo.get('status', '').lower() in ['done', 'completed'] else 'No',
                'NOTES': todo.get('notes', '')
            }
            converted['TO-DO REVIEW'].append(todo_item)
        print(f"Converted {len(converted['TO-DO REVIEW'])} todo reviews to TO-DO REVIEW")
    
    # Map headlines
    if 'headlines' in data:
        converted['HEADLINES'] = []
        for headline in data['headlines']:
            if isinstance(headline, dict):
                converted['HEADLINES'].append(headline.get('text', str(headline)))
            else:
                converted['HEADLINES'].append(str(headline))
        print(f"Converted {len(converted['HEADLINES'])} headlines")
    
    # Copy other fields that might be present
    for key in ['MEETING RATING', 'average_rating', 'meeting_date', 'attendees']:
        if key in data:
            converted[key] = data[key]
    
    print(f"Final converted keys: {list(converted.keys())}")
    return converted
def parse_l10_text(text):
    """Parse the structured L10 text output into a dictionary format"""
    sections = {
        'HEADLINES': [],
        'TO-DO REVIEW': [],
        'SCORECARD/METRICS REVIEW': '',
        'ROCK REVIEW': '',
        'CUSTOMER/EMPLOYEE HEADLINES': [],
        'ISSUES LIST (IDS)': [],
        'NEW TO-DOS': [],
        'CASCADING MESSAGES': '',
        'MEETING RATING': []
    }
    
    current_section = None
    current_item = {}
    lines = text.strip().split('\n')
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        if not line:
            i += 1
            continue
            
        if line.startswith('**') and line.endswith('**'):
            if current_item and current_section in ['TO-DO REVIEW', 'ISSUES LIST (IDS)', 'NEW TO-DOS']:
                sections[current_section].append(current_item)
                current_item = {}
            
            current_section = line.strip('*').strip()
            i += 1
            continue
        
        # Handle different sections
        if current_section == 'HEADLINES':
            if line.startswith('-'):
                sections['HEADLINES'].append(line[1:].strip())
        
        elif current_section == 'TO-DO REVIEW':
            if line == '---':
                if current_item:
                    sections['TO-DO REVIEW'].append(current_item)
                    current_item = {}
            elif line.startswith('WHO:'):
                current_item['WHO'] = line[4:].strip()
            elif line.startswith('TO-DO:'):
                current_item['TO-DO'] = line[6:].strip()
            elif line.startswith('DONE?:'):
                current_item['DONE?'] = line[6:].strip()
            elif line.startswith('NOTES:'):
                current_item['NOTES'] = line[6:].strip()
        
        elif current_section == 'ISSUES LIST (IDS)':
            if line == '---':
                if current_item:
                    sections['ISSUES LIST (IDS)'].append(current_item)
                    current_item = {}
            elif line.startswith('ISSUE:'):
                current_item['issue'] = line[6:].strip()
            elif line.startswith('RAISED BY:'):
                current_item['raised_by'] = line[10:].strip()
            elif line.startswith('DISCUSSION:'):
                current_item['discussion'] = line[11:].strip()
        
        elif current_section == 'NEW TO-DOS':
            if line == '---':
                if current_item:
                    sections['NEW TO-DOS'].append(current_item)
                    current_item = {}
            elif line.startswith('WHO:'):
                current_item['WHO'] = line[4:].strip()
            elif line.startswith('TO-DO:'):
                current_item['TO-DO'] = line[6:].strip()
            elif line.startswith('DUE:'):
                current_item['DUE'] = line[4:].strip()
        
        elif current_section == 'CUSTOMER/EMPLOYEE HEADLINES':
            if line.startswith('-'):
                sections['CUSTOMER/EMPLOYEE HEADLINES'].append(line[1:].strip())
            elif line != 'None discussed' and line != '---':
                sections['CUSTOMER/EMPLOYEE HEADLINES'].append(line)
        
        elif current_section == 'MEETING RATING':
            if ':' in line and not line.startswith('Average'):
                name, rating = line.split(':', 1)
                sections['MEETING RATING'].append({
                    'name': name.strip(),
                    'rating': rating.strip()
                })
            elif line.startswith('Average:'):
                sections['average_rating'] = line.split(':')[1].strip()
        
        i += 1
    
    if current_item and current_section in ['TO-DO REVIEW', 'ISSUES LIST (IDS)', 'NEW TO-DOS']:
        sections[current_section].append(current_item)
    
    return sections

def find_section_row(ws, keywords, start_row=1, end_row=None):
    """Find row containing any of the keywords"""
    if end_row is None:
        end_row = min(30, ws.max_row)
    
    for row in range(start_row, end_row + 1):
        for col in range(1, min(7, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).upper()
                for keyword in keywords:
                    if keyword.upper() in cell_str:
                        return row
    return None

def copy_row_format(ws, source_row, target_row):
    """Copy formatting from source row to target row"""
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

def populate_l10_from_text(text_input, template_path, output_path):
    """Original function to populate L10 template from structured text input"""
    
    # Parse the text input
    data = parse_l10_text(text_input)
    
    # Load Excel template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    print("=== POPULATING TEMPLATE ===")
    print("Finding template sections...")
    
    # Find key sections in the template
    headlines_row = find_section_row(ws, ['Headlines:', 'HEADLINE'])
    good_news_row = find_section_row(ws, ['Good News'], start_row=headlines_row or 1)
    todo_header_row = find_section_row(ws, ['To-Do List', 'TO-DO'], start_row=5)
    issues_row = find_section_row(ws, ['Issues (IDS)', 'ISSUES'], start_row=8)
    rating_row = find_section_row(ws, ['Did we start/end', 'RATING'], start_row=12)
    
    print(f"Found sections at rows - Headlines: {headlines_row}, TO-DO: {todo_header_row}, Issues: {issues_row}, Rating: {rating_row}")
    
    # Track how many rows we've inserted
    total_inserted = 0
    
    # 1. HEADLINES Section - Place in the Good News area
    if data['HEADLINES'] and good_news_row:
        print(f"Processing {len(data['HEADLINES'])} headlines...")
        col = 2  # Start at column B
        for i, headline in enumerate(data['HEADLINES']):
            if col <= 6:  # Up to column F
                ws.cell(row=good_news_row, column=col, value=headline)
                col += 1
    
    # 2. TO-DO REVIEW Section
    if data['TO-DO REVIEW'] and todo_header_row:
        print(f"Processing {len(data['TO-DO REVIEW'])} TO-DO items...")
        insert_row = todo_header_row + 1
        
        for r in range(todo_header_row, todo_header_row + 5):
            if ws.cell(row=r, column=1).value and 'WHO' in str(ws.cell(row=r, column=1).value).upper():
                insert_row = r + 1
                break
        
        insert_row += total_inserted
        
        num_todos = len(data['TO-DO REVIEW'])
        ws.insert_rows(insert_row, num_todos)
        total_inserted += num_todos
        
        for i, todo in enumerate(data['TO-DO REVIEW']):
            row = insert_row + i
            ws.cell(row=row, column=1, value=todo.get('WHO', ''))
            ws.cell(row=row, column=2, value=todo.get('TO-DO', ''))
            ws.cell(row=row, column=3, value=todo.get('DONE?', ''))
            ws.cell(row=row, column=4, value=todo.get('NOTES', ''))
            if insert_row > 1:
                copy_row_format(ws, insert_row - 1, row)
    
    # 3. ISSUES LIST Section
    if data['ISSUES LIST (IDS)'] and issues_row:
        print(f"Processing {len(data['ISSUES LIST (IDS)'])} issues...")
        insert_row = issues_row + 1 + total_inserted
        
        num_issues = len(data['ISSUES LIST (IDS)'])
        ws.insert_rows(insert_row, num_issues)
        total_inserted += num_issues
        
        for i, issue in enumerate(data['ISSUES LIST (IDS)']):
            row = insert_row + i
            issue_text = f"{issue.get('issue', '')} - {issue.get('raised_by', '')} - {issue.get('discussion', '')}"
            ws.cell(row=row, column=2, value=issue_text)
            copy_row_format(ws, issues_row + total_inserted - num_issues, row)
    
    # 4. NEW TO-DOS Section
    if data['NEW TO-DOS'] and todo_header_row:
        print(f"Processing {len(data['NEW TO-DOS'])} new TO-DOs...")
        insert_row = (todo_header_row + total_inserted + 
                     len(data['TO-DO REVIEW']) + 2)
        
        ws.insert_rows(insert_row, 1)
        ws.cell(row=insert_row, column=1, value="NEW ACTION ITEMS THIS WEEK:")
        ws.cell(row=insert_row, column=1).font = Font(bold=True)
        total_inserted += 1
        
        num_new_todos = len(data['NEW TO-DOS'])
        ws.insert_rows(insert_row + 1, num_new_todos)
        total_inserted += num_new_todos
        
        for i, todo in enumerate(data['NEW TO-DOS']):
            row = insert_row + 1 + i
            ws.cell(row=row, column=1, value=todo.get('WHO', ''))
            ws.cell(row=row, column=2, value=todo.get('TO-DO', ''))
            ws.cell(row=row, column=3, value=todo.get('DUE', ''))
    
    # 5. MEETING RATING Section
    if data['MEETING RATING'] and rating_row:
        print(f"Processing {len(data['MEETING RATING'])} ratings...")
        insert_row = rating_row + 1 + total_inserted
        
        num_ratings = len(data['MEETING RATING']) + 2
        ws.insert_rows(insert_row, num_ratings)
        
        ws.cell(row=insert_row, column=1, value="Meeting Ratings:")
        ws.cell(row=insert_row, column=1).font = Font(bold=True)
        
        for i, rating in enumerate(data['MEETING RATING']):
            row = insert_row + 1 + i
            ws.cell(row=row, column=1, value=rating['name'])
            ws.cell(row=row, column=2, value=f"{rating['rating']}/10")
        
        if 'average_rating' in data:
            avg_row = insert_row + len(data['MEETING RATING']) + 1
            ws.cell(row=avg_row, column=1, value="Average:")
            ws.cell(row=avg_row, column=2, value=f"{data['average_rating']}/10")
            ws.cell(row=avg_row, column=1).font = Font(bold=True)
    
    # Save the file
    wb.save(output_path)
    print(f"\nSuccessfully saved populated L10 to: {output_path}")
    return output_path

# Original test function
def test_with_text_input():
    """Test the population with text input from file"""
    
    if os.path.exists('l10_output.txt'):
        with open('l10_output.txt', 'r') as f:
            text_input = f.read()
        print("Using l10_output.txt file...")
    else:
        print("l10_output.txt not found. Using sample data...")
        text_input = """[sample data]"""
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f'populated_l10_{timestamp}.xlsx'
    
    populate_l10_from_text(
        text_input,
        'L10 Summary Template 1.xlsx',
        output_filename
    )
    
    print(f"Output saved as: {output_filename}")


# NEW: Enhanced L10Processor class with automation features
class L10Processor:
    def __init__(self):
        self.sections = {
            'HEADLINES': [],
            'TO-DO REVIEW': [],
            'SCORECARD/METRICS REVIEW': '',
            'ROCK REVIEW': '',
            'CUSTOMER/EMPLOYEE HEADLINES': [],
            'ISSUES LIST (IDS)': [],
            'NEW TO-DOS': [],
            'AI IDENTIFIED ITEMS': [],  # New section for AI items
            'CASCADING MESSAGES': '',
            'MEETING RATING': []
        }
    
    def parse_l10_text(self, text):
        """Parse text - reuses the original function"""
        return parse_l10_text(text)
    
    def duplicate_previous_sheet(self, previous_path, output_path, next_meeting_date=None):
        """
        Duplicate the previous L10 sheet as per the meeting process.
        This mimics the "move or copy" -> "move to end" -> "create a copy" workflow
        """
        # Load the previous workbook
        wb = openpyxl.load_workbook(previous_path)
        ws = wb.active
        
        # Update the date if provided
        if next_meeting_date:
            # Look for date in typical locations (usually in header area)
            for row in range(1, 5):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        # Look for date patterns
                        if re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', str(cell.value)):
                            cell.value = next_meeting_date
                            break
        
        # Save as new file
        wb.save(output_path)
        return wb, ws
    
    def find_existing_todos(self, ws):
        """Extract existing TO-DOs from the worksheet"""
        existing_todos = []
        
        # Find TO-DO section
        todo_row = find_section_row(ws, ['To-Do List', 'TO-DO'])
        
        if todo_row:
            # Look for TO-DO items after the header
            for row in range(todo_row + 1, ws.max_row + 1):
                who = ws.cell(row=row, column=1).value
                todo = ws.cell(row=row, column=2).value
                done = ws.cell(row=row, column=3).value
                
                if who and todo:
                    existing_todos.append({
                        'WHO': str(who).strip(),
                        'TO-DO': str(todo).strip(),
                        'DONE?': str(done).strip() if done else '',
                        'row': row
                    })
                elif not who and not todo and row > todo_row + 5:
                    # Likely end of TO-DO section
                    break
        
        return existing_todos
    
    def compare_todos(self, new_todos, existing_todos):
        """
        Compare new TO-DOs with existing ones to avoid duplicates.
        Returns truly new items and items that need updates.
        """
        truly_new = []
        updates = []
        
        for new_todo in new_todos:
            found = False
            for existing in existing_todos:
                # Check if it's the same TO-DO (by WHO and TO-DO text)
                if (new_todo.get('WHO', '').lower() == existing.get('WHO', '').lower() and
                    new_todo.get('TO-DO', '').lower() == existing.get('TO-DO', '').lower()):
                    found = True
                    # Check if status needs update
                    if new_todo.get('DONE?', '') != existing.get('DONE?', ''):
                        updates.append({
                            'row': existing['row'],
                            'new_status': new_todo.get('DONE?', ''),
                            'new_notes': new_todo.get('NOTES', '')
                        })
                    break
            
            if not found:
                truly_new.append(new_todo)
        
        return truly_new, updates
    
    def add_ai_section(self, ws, ai_items, start_row):
        """Add a dedicated AI Identified Items section"""
        # Add header
        ws.cell(row=start_row, column=1, value="AI IDENTIFIED ITEMS (Review & Move to Appropriate Sections)")
        ws.cell(row=start_row, column=1).font = Font(bold=True, color="0066CC")
        
        # Add border
        border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        current_row = start_row + 1
        
        # Add TO-DOs identified by AI
        if 'new_todos' in ai_items and ai_items['new_todos']:
            ws.cell(row=current_row, column=1, value="Potential TO-DOs:")
            ws.cell(row=current_row, column=1).font = Font(italic=True)
            current_row += 1
            
            for todo in ai_items['new_todos']:
                ws.cell(row=current_row, column=1, value=f"• {todo.get('WHO', 'TBD')}")
                ws.cell(row=current_row, column=2, value=todo.get('TO-DO', ''))
                ws.cell(row=current_row, column=3, value=todo.get('DUE', ''))
                current_row += 1
        
        # Add Issues identified by AI
        if 'new_issues' in ai_items and ai_items['new_issues']:
            current_row += 1
            ws.cell(row=current_row, column=1, value="Potential Issues:")
            ws.cell(row=current_row, column=1).font = Font(italic=True)
            current_row += 1
            
            for issue in ai_items['new_issues']:
                ws.cell(row=current_row, column=1, value=f"• {issue.get('issue', '')}")
                ws.cell(row=current_row, column=2, value=f"Raised by: {issue.get('raised_by', 'TBD')}")
                current_row += 1
        
        return current_row
    
    def calculate_next_meeting_date(self, cadence='weekly', last_date=None):
        """Calculate the next meeting date based on cadence"""
        if last_date:
            if isinstance(last_date, str):
                # Parse the date string
                for fmt in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
                    try:
                        last_date = datetime.strptime(last_date, fmt)
                        break
                    except:
                        continue
        
        if not last_date:
            last_date = datetime.now()
        
        if cadence == 'weekly':
            next_date = last_date + timedelta(days=7)
        elif cadence == 'biweekly':
            next_date = last_date + timedelta(days=14)
        else:
            # Default to weekly
            next_date = last_date + timedelta(days=7)
        
        return next_date.strftime('%m/%d/%Y')
    
    def process_l10_automation(self, previous_template_path, new_data_path, output_path, 
                              meeting_cadence='weekly', last_meeting_date=None):
        """
        Main automation function that implements the L10 meeting workflow
        """
        print("=== L10 MEETING AUTOMATION ===")
        
        # Step 1: Calculate next meeting date
        next_meeting_date = self.calculate_next_meeting_date(meeting_cadence, last_meeting_date)
        print(f"Next meeting date: {next_meeting_date}")
        
        # Step 2: Duplicate previous sheet
        print("Duplicating previous L10 sheet...")
        wb, ws = self.duplicate_previous_sheet(previous_template_path, output_path, next_meeting_date)
        
        # Step 3: Parse new meeting data
        if isinstance(new_data_path, str) and new_data_path.endswith('.txt'):
            with open(new_data_path, 'r') as f:
                text_input = f.read()
            new_data = self.parse_l10_text(text_input)
        else:
            # Assume it's already parsed data
            new_data = new_data_path
        
        # Step 4: Find existing TO-DOs
        print("Analyzing existing TO-DOs...")
        existing_todos = self.find_existing_todos(ws)
        print(f"Found {len(existing_todos)} existing TO-DOs")
        
        # Step 5: Compare and identify truly new items
        all_new_todos = new_data.get('NEW TO-DOS', []) + new_data.get('TO-DO REVIEW', [])
        truly_new_todos, todo_updates = self.compare_todos(all_new_todos, existing_todos)
        print(f"Identified {len(truly_new_todos)} new TO-DOs")
        
        # Step 6: Update existing TO-DOs status (but don't remove completed ones)
        for update in todo_updates:
            ws.cell(row=update['row'], column=3, value=update['new_status'])
            if update.get('new_notes'):
                ws.cell(row=update['row'], column=4, value=update['new_notes'])
        
        # Step 7: Add AI Identified Items section
        ai_items = {
            'new_todos': truly_new_todos,
            'new_issues': new_data.get('ISSUES LIST (IDS)', [])
        }
        
        # Find a good place to add AI section (after existing content)
        last_content_row = ws.max_row
        ai_section_start = last_content_row + 2
        
        self.add_ai_section(ws, ai_items, ai_section_start)
        
        # Step 8: Update headlines if any
        if new_data.get('HEADLINES'):
            headlines_row = find_section_row(ws, ['Headlines:', 'HEADLINE'])
            if headlines_row:
                good_news_row = find_section_row(ws, ['Good News'], start_row=headlines_row or 1)
                if good_news_row:
                    col = 2
                    for headline in new_data['HEADLINES'][:5]:  # Max 5 headlines
                        if col <= 6:
                            ws.cell(row=good_news_row, column=col, value=headline)
                            col += 1
        
        # Step 9: Save the updated workbook
        wb.save(output_path)
        print(f"Successfully saved automated L10 to: {output_path}")
        
        return {
            'output_path': output_path,
            'next_meeting_date': next_meeting_date,
            'new_todos_count': len(truly_new_todos),
            'updated_todos_count': len(todo_updates),
            'ai_items_added': len(truly_new_todos) + len(new_data.get('ISSUES LIST (IDS)', []))
        }


# Keep backward compatibility
if __name__ == "__main__":
    test_with_text_input()
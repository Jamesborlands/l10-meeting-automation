#!/usr/bin/env python3
"""
Check the specific Excel file you're looking at
"""

import openpyxl

def check_file_details(filename):
    """Check detailed file information"""
    print(f"📁 CHECKING FILE: {filename}")
    print("="*50)
    
    try:
        wb = openpyxl.load_workbook(filename)
        
        # List all sheets
        print(f"All sheets ({len(wb.sheetnames)}):")
        for i, sheet_name in enumerate(wb.sheetnames):
            if i >= len(wb.sheetnames) - 3:  # Show last 3 sheets
                print(f"  {i+1}. {sheet_name} {'<-- LATEST' if i == len(wb.sheetnames) - 1 else ''}")
        
        # Check the latest sheet
        latest_sheet = wb[wb.sheetnames[-1]]
        print(f"\nLatest sheet details:")
        print(f"  Name: {latest_sheet.title}")
        print(f"  Max row: {latest_sheet.max_row}")
        print(f"  Max col: {latest_sheet.max_column}")
        
        # Check AI section area specifically
        print(f"\nChecking rows 50-65 for AI content:")
        for row in range(50, min(66, latest_sheet.max_row + 1)):
            row_content = []
            for col in range(1, 6):
                cell = latest_sheet.cell(row=row, column=col)
                value = cell.value
                if value:
                    # Check if cell has any special formatting
                    row_height = latest_sheet.row_dimensions[row].height if row in latest_sheet.row_dimensions else "default"
                    col_width = latest_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width if openpyxl.utils.get_column_letter(col) in latest_sheet.column_dimensions else "default"
                    
                    row_content.append(f"C{col}:'{str(value)[:50]}{'...' if len(str(value)) > 50 else ''}'")
            
            if row_content:
                row_height = latest_sheet.row_dimensions[row].height if row in latest_sheet.row_dimensions else None
                height_info = f" [H:{row_height}]" if row_height else ""
                print(f"  Row {row}{height_info}: {' | '.join(row_content)}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

# Check the file you're most likely looking at
files_to_check = [
    'debug_render_output.xlsx',
    'test_render_output.xlsx', 
    'test_local_output.xlsx'
]

for filename in files_to_check:
    try:
        check_file_details(filename)
        print("\n" + "="*70 + "\n")
    except FileNotFoundError:
        print(f"❌ File not found: {filename}\n")
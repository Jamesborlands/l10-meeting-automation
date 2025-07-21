import openpyxl
import os

print("Testing Excel files...\n")

# List all Excel files
excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
print(f"Found {len(excel_files)} Excel files:\n")

for file in excel_files:
    try:
        # Get file size
        size = os.path.getsize(file)
        print(f"File: {file}")
        print(f"  Size: {size} bytes")
        
        # Try to open it
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        print(f"  ✓ Valid Excel file")
        print(f"  Sheet name: {ws.title}")
        print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
    except Exception as e:
        print(f"  ✗ Error: {str(e)}")
    
    print()

# Special check for the template
template = "L10 Summary Template 1.xlsx"
if os.path.exists(template):
    print(f"\nChecking {template} specifically:")
    # Try reading first few bytes
    with open(template, 'rb') as f:
        header = f.read(4)
        if header == b'PK\x03\x04':
            print("  File has correct ZIP header (Excel format)")
        else:
            print(f"  File has incorrect header: {header}")
            print("  This file might be corrupted or not a real Excel file")
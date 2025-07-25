#!/usr/bin/env python3
"""
Investigate what's actually on Render right now
"""

import requests
import json

def check_render_current_state():
    """Check what version and code is actually running on Render"""
    print("🔍 INVESTIGATING RENDER LIVE STATE")
    print("="*60)
    
    # Test the health endpoint
    try:
        health_response = requests.get('https://l10-meeting-automation-29fl.onrender.com/health', timeout=10)
        print(f"Health endpoint: {health_response.status_code}")
        if health_response.status_code == 200:
            health_data = health_response.json()
            print(f"Health response: {health_data}")
            
            # Check if version field exists (we added this in recent commits)
            if 'version' in health_data:
                print("✅ Render has recent code (version field present)")
            else:
                print("❌ Render might have old code (no version field)")
        
    except Exception as e:
        print(f"❌ Health check failed: {e}")
    
    # Test the debug endpoint
    try:
        debug_response = requests.get('https://l10-meeting-automation-29fl.onrender.com/debug', timeout=10)
        print(f"\nDebug endpoint: {debug_response.status_code}")
        if debug_response.status_code == 200:
            debug_data = debug_response.json()
            files = debug_data.get('files', [])
            print(f"Files on Render: {len(files)} files")
            
            # Check for files that should/shouldn't be there
            expected_files = ['l10_processor.py', 'app.py', 'l10_sheet_automation.py']
            old_debug_files = ['debug_ai_section.py', 'fix_ai_section.py', 'test_ai_section.py']
            
            for file in expected_files:
                if file in files:
                    print(f"✅ {file} present")
                else:
                    print(f"❌ {file} MISSING")
            
            old_files_present = [f for f in old_debug_files if f in files]
            if old_files_present:
                print(f"⚠️  Old debug files still present: {old_files_present}")
            else:
                print("✅ Old debug files cleaned up")
        
    except Exception as e:
        print(f"❌ Debug check failed: {e}")
    
    # Test if the test-conversion endpoint exists (we added this recently)
    try:
        test_data = {
            "meeting_data": {
                "new_commitments": [{"who": "Test", "task": "Test task", "due_date": "Test"}]
            }
        }
        
        conversion_response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/test-conversion',
            json=test_data,
            timeout=15
        )
        
        print(f"\nTest-conversion endpoint: {conversion_response.status_code}")
        
        if conversion_response.status_code == 200:
            print("✅ Test-conversion endpoint exists (recent code)")
            result = conversion_response.json()
            if result.get('conversion_successful'):
                print("✅ Conversion function is working on Render!")
            else:
                print("❌ Conversion function exists but not working properly")
                print(f"Conversion result: {result}")
        
        elif conversion_response.status_code == 404:
            print("❌ Test-conversion endpoint missing (old code)")
        else:
            print(f"❌ Unexpected response: {conversion_response.text}")
        
    except Exception as e:
        print(f"❌ Conversion test failed: {e}")

def test_actual_conversion():
    """Test conversion with your actual data"""
    print(f"\n🧪 TESTING WITH YOUR ACTUAL DATA")
    print("="*60)
    
    your_data = {
        "meeting_date": "2023-10-05",
        "attendees": ["Josh Keshen", "Lyndsey Dunnavant", "Josh Weinberg", "Regan Gentry", "Lisa Schick"],
        "new_commitments": [
            {
                "who": "Lyndsey Dunnavant",
                "task": "Ensure training policy is revised and prepared for compliance review.",
                "due_date": "Next meeting",
                "context": "To align with regulatory standards.",
                "dependencies": "Gather input from team and maintain policy visibility."
            },
            {
                "who": "Josh Weinberg", 
                "task": "Monitor updates on audits and ensure compliance documentation is complete.",
                "due_date": "Next meeting",
                "context": "To facilitate smooth audits in future and ensure all states are compliant.",
                "dependencies": "Collaboration with state teams."
            }
        ],
        "issues_discussed": [
            {
                "issue": "The need to adjust the training policy.",
                "raised_by": "Lyndsey Dunnavant",
                "context": "The existing training policy is not being adhered to, and regulators may question training compliance.",
                "decision": "Lyndsey will redline the training policy for revision.",
                "owner": "Lyndsey Dunnavant"
            },
            {
                "issue": "Upcoming reports and responsibilities for state audits.",
                "raised_by": "Josh Weinberg",
                "context": "Kansas is leveraging Washington's audits, and Mississippi's audits are pending.",
                "decision": "Keep monitoring the situation with state audits; further updates are needed.",
                "owner": "Josh Weinberg"
            }
        ]
    }
    
    payload = {"meeting_data": your_data, "excel_url": ""}
    
    try:
        # Test the main processing endpoint
        response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload,
            timeout=60
        )
        
        print(f"Process-l10 response: {response.status_code}")
        
        if response.status_code == 200:
            # Save and analyze the Excel file
            with open('live_render_test.xlsx', 'wb') as f:
                f.write(response.content)
            
            print(f"✅ Excel file created: {len(response.content)} bytes")
            
            # Quick analysis
            import openpyxl
            wb = openpyxl.load_workbook('live_render_test.xlsx')
            latest_sheet = wb[wb.sheetnames[-1]]
            
            print(f"✅ Sheet: {latest_sheet.title}, Max row: {latest_sheet.max_row}")
            
            # Look for AI section content
            ai_items_found = 0
            ai_row = None
            
            for row in range(45, min(70, latest_sheet.max_row + 1)):
                cell_value = latest_sheet.cell(row=row, column=1).value
                if cell_value and "AI IDENTIFIED" in str(cell_value):
                    ai_row = row
                    print(f"✅ AI section found at row {row}")
                    
                    # Count actual items
                    for check_row in range(row + 1, min(row + 15, latest_sheet.max_row + 1)):
                        has_content = False
                        for col in range(1, 6):
                            val = latest_sheet.cell(row=check_row, column=col).value
                            if val and str(val).strip() and "Potential" not in str(val) and "WHO" not in str(val) and "RAISED BY" not in str(val):
                                has_content = True
                                break
                        if has_content:
                            ai_items_found += 1
                    break
            
            if ai_row:
                print(f"✅ AI section exists, Items found: {ai_items_found}")
                if ai_items_found > 0:
                    print("🎉 AI SECTION IS WORKING!")
                else:
                    print("❌ AI section empty - conversion not working")
            else:
                print("❌ No AI section found")
            
            wb.close()
        
        else:
            print(f"❌ Process failed: {response.status_code}")
            print(response.text)
    
    except Exception as e:
        print(f"❌ Processing test failed: {e}")

if __name__ == "__main__":
    check_render_current_state()
    test_actual_conversion()
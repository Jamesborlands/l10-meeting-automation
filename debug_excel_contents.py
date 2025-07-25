#!/usr/bin/env python3
"""
Debug script to examine the actual contents of the Excel files
"""

import openpyxl
import json

def debug_excel_file(filename):
    """Debug the Excel file contents around the AI section"""
    print(f"🔍 DEBUGGING: {filename}")
    print("="*60)
    
    try:
        wb = openpyxl.load_workbook(filename)
        sheets = wb.sheetnames
        latest_sheet = wb[sheets[-1]]
        
        print(f"Latest sheet: {latest_sheet.title}")
        print(f"Max row: {latest_sheet.max_row}, Max col: {latest_sheet.max_column}")
        
        # Find AI section
        ai_section_row = None
        for row in range(1, latest_sheet.max_row + 1):
            for col in range(1, latest_sheet.max_column + 1):
                cell_value = latest_sheet.cell(row=row, column=col).value
                if cell_value and "AI IDENTIFIED ITEMS" in str(cell_value):
                    ai_section_row = row
                    break
            if ai_section_row:
                break
        
        if ai_section_row:
            print(f"AI section found at row: {ai_section_row}")
            
            # Print the next 15 rows after AI section
            print("\nRows after AI section:")
            for row in range(ai_section_row, min(ai_section_row + 15, latest_sheet.max_row + 1)):
                row_data = []
                for col in range(1, 8):  # Check first 7 columns
                    cell_value = latest_sheet.cell(row=row, column=col).value
                    if cell_value:
                        row_data.append(f"Col{col}: '{cell_value}'")
                
                if row_data:
                    print(f"  Row {row}: {' | '.join(row_data)}")
                else:
                    print(f"  Row {row}: EMPTY")
        else:
            print("❌ AI section not found!")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

def test_render_with_debug():
    """Test Render with detailed debug output"""
    import requests
    
    # Load sample data
    with open('sample_l10_data.json', 'r') as f:
        sample_data = json.load(f)
    
    print("🧪 TESTING RENDER WITH DEBUG")
    print("="*60)
    print(f"Sample data keys: {list(sample_data.keys())}")
    print(f"NEW TO-DOS: {len(sample_data.get('NEW TO-DOS', []))}")
    print(f"ISSUES LIST: {len(sample_data.get('ISSUES LIST (IDS)', []))}")
    
    if sample_data.get('NEW TO-DOS'):
        print("NEW TO-DOS content:")
        for i, todo in enumerate(sample_data['NEW TO-DOS']):
            print(f"  {i+1}. {todo}")
    
    if sample_data.get('ISSUES LIST (IDS)'):
        print("ISSUES LIST content:")
        for i, issue in enumerate(sample_data['ISSUES LIST (IDS)']):
            print(f"  {i+1}. {issue}")
    
    # Send to Render
    payload = {
        "meeting_data": sample_data,
        "excel_url": ""
    }
    
    try:
        response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload,
            timeout=60
        )
        
        if response.status_code == 200:
            with open('debug_render_output.xlsx', 'wb') as f:
                f.write(response.content)
            print("✅ Response received, saved as debug_render_output.xlsx")
            
            # Debug the output
            debug_excel_file('debug_render_output.xlsx')
        else:
            print(f"❌ Error response: {response.status_code}")
            print(response.text)
            
    except Exception as e:
        print(f"❌ Request error: {e}")

if __name__ == "__main__":
    # First debug existing files
    files = ['test_local_output.xlsx', 'test_render_output.xlsx']
    for filename in files:
        try:
            debug_excel_file(filename)
            print("\n")
        except FileNotFoundError:
            print(f"File not found: {filename}\n")
    
    # Then test Render with fresh debug
    test_render_with_debug()
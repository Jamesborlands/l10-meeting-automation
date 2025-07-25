#!/usr/bin/env python3
"""
Test Render with your actual meeting data
"""

import requests
import json

# Your actual meeting data
your_meeting_data = {
  "meeting_date": "2023-10-05",
  "attendees": [
    "Josh Keshen",
    "Lyndsey Dunnavant",
    "Josh Weinberg", 
    "Regan Gentry",
    "Lisa Schick"
  ],
  "headlines": [
    {
      "text": "Removal of training suggestion to archive the training policy and updates on new state audits."
    }
  ],
  "todo_review": [
    {
      "who": "Lyndsey Dunnavant",
      "todo": "Redline the training policy to include ongoing training references.",
      "status": "Not Done",
      "notes": "Lyndsey believes it is essential to retain a training policy to avoid regulatory issues.",
      "original_due": "Next meeting"
    }
  ],
  "issues_discussed": [
    {
      "issue": "The need to adjust the training policy.",
      "raised_by": "Lyndsey Dunnavant",
      "context": "The existing training policy is not being adhered to, and regulators may question training compliance.",
      "discussion_points": [
        "Lyndsey noted the current policy mentions comprehensive training that is not being provided.",
        "Josh expressed concern about removing training references completely.",
        "Josh suggested revising the training policy to include ongoing training versus a fixed schedule."
      ],
      "decision": "Lyndsey will redline the training policy for revision.",
      "owner": "Lyndsey Dunnavant"
    },
    {
      "issue": "Upcoming reports and responsibilities for state audits.",
      "raised_by": "Josh Weinberg",
      "context": "Kansas is leveraging Washington's audits, and Mississippi's audits are pending.",
      "discussion_points": [
        "Josh provided updates on audits in different states.",
        "Lyndsey mentioned Mississippi processing and possible split of work."
      ],
      "decision": "Keep monitoring the situation with state audits; further updates are needed.",
      "owner": "Josh Weinberg"
    }
  ],
  "new_commitments": [
    {
      "who": "Lyndsey Dunnavant",
      "task": "Ensure training policy is revised and prepared for compliance review.",
      "due_date": "Next meeting",
      "context": "To align with regulatory standards.",
      "dependencies": "Gather input from team and maintain policy visibility."
    },
    {
      "who": "Josh Weinberg",
      "task": "Monitor updates on audits and ensure compliance documentation is complete.",
      "due_date": "Next meeting",
      "context": "To facilitate smooth audits in future and ensure all states are compliant.",
      "dependencies": "Collaboration with state teams."
    }
  ]
}

def test_your_data():
    """Test Render with your meeting data"""
    print("🧪 Testing Render with your meeting data")
    print("="*60)
    
    payload = {
        "meeting_data": your_meeting_data,
        "excel_url": ""
    }
    
    print(f"Sending data with:")
    print(f"  - {len(your_meeting_data.get('new_commitments', []))} new commitments")
    print(f"  - {len(your_meeting_data.get('issues_discussed', []))} issues discussed")
    print(f"  - {len(your_meeting_data.get('todo_review', []))} todo reviews")
    
    try:
        response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload,
            timeout=60
        )
        
        print(f"\nResponse status: {response.status_code}")
        
        if response.status_code == 200:
            # Save the file
            with open('your_data_output.xlsx', 'wb') as f:
                f.write(response.content)
            
            print(f"✅ Success! File saved as: your_data_output.xlsx")
            print(f"✅ File size: {len(response.content)} bytes")
            
            # Quick check of the Excel file
            import openpyxl
            wb = openpyxl.load_workbook('your_data_output.xlsx') 
            latest_sheet = wb[wb.sheetnames[-1]]
            
            print(f"✅ New sheet created: {latest_sheet.title}")
            print(f"✅ Sheet dimensions: {latest_sheet.max_row} rows")
            
            # Look for AI section
            for row in range(50, min(70, latest_sheet.max_row + 1)):
                cell_value = latest_sheet.cell(row=row, column=1).value
                if cell_value and "AI IDENTIFIED" in str(cell_value):
                    print(f"✅ AI section found at row {row}")
                    
                    # Count items
                    item_count = 0
                    for check_row in range(row + 1, min(row + 20, latest_sheet.max_row + 1)):
                        check_value = latest_sheet.cell(row=check_row, column=2).value
                        if check_value and str(check_value).strip():
                            item_count += 1
                    
                    print(f"✅ Found approximately {item_count} items in AI section")
                    break
            
            wb.close()
            
        else:
            print(f"❌ Error: {response.status_code}")
            print(response.text)
    
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    test_your_data()
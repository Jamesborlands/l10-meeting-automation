#!/usr/bin/env python3
"""
Test the conversion locally with your data
"""

import json
import shutil
from l10_sheet_automation import L10SheetAutomation
from l10_processor import parse_l10_json

your_meeting_data = {
  "meeting_date": "2023-10-05",
  "attendees": ["Josh Keshen", "Lyndsey Dunnavant", "Josh Weinberg", "Regan Gentry", "Lisa Schick"],
  "headlines": [{"text": "Removal of training suggestion to archive the training policy and updates on new state audits."}],
  "todo_review": [{
    "who": "Lyndsey Dunnavant",
    "todo": "Redline the training policy to include ongoing training references.",
    "status": "Not Done",
    "notes": "Lyndsey believes it is essential to retain a training policy to avoid regulatory issues.",
    "original_due": "Next meeting"
  }],
  "issues_discussed": [
    {
      "issue": "The need to adjust the training policy.",
      "raised_by": "Lyndsey Dunnavant",
      "context": "The existing training policy is not being adhered to, and regulators may question training compliance.",
      "discussion_points": ["Lyndsey noted the current policy mentions comprehensive training that is not being provided.", "Josh expressed concern about removing training references completely.", "Josh suggested revising the training policy to include ongoing training versus a fixed schedule."],
      "decision": "Lyndsey will redline the training policy for revision.",
      "owner": "Lyndsey Dunnavant"
    },
    {
      "issue": "Upcoming reports and responsibilities for state audits.",
      "raised_by": "Josh Weinberg", 
      "context": "Kansas is leveraging Washington's audits, and Mississippi's audits are pending.",
      "discussion_points": ["Josh provided updates on audits in different states.", "Lyndsey mentioned Mississippi processing and possible split of work."],
      "decision": "Keep monitoring the situation with state audits; further updates are needed.",
      "owner": "Josh Weinberg"
    }
  ],
  "new_commitments": [
    {
      "who": "Lyndsey Dunnavant",
      "task": "Ensure training policy is revised and prepared for compliance review.",
      "due_date": "Next meeting",
      "context": "To align with regulatory standards.",
      "dependencies": "Gather input from team and maintain policy visibility."
    },
    {
      "who": "Josh Weinberg",
      "task": "Monitor updates on audits and ensure compliance documentation is complete.",
      "due_date": "Next meeting",
      "context": "To facilitate smooth audits in future and ensure all states are compliant.",
      "dependencies": "Collaboration with state teams."
    }
  ]
}

def test_local_conversion():
    """Test the conversion and automation locally"""
    print("🧪 Testing local conversion with your data")
    print("="*60)
    
    # Parse the data
    parsed_data = parse_l10_json(your_meeting_data)
    
    print(f"Parsed data keys: {list(parsed_data.keys())}")
    print(f"NEW TO-DOS: {len(parsed_data.get('NEW TO-DOS', []))}")
    print(f"ISSUES LIST: {len(parsed_data.get('ISSUES LIST (IDS)', []))}")
    
    # Create test workbook
    shutil.copy('L10 Summary Template 1.xlsx', 'local_conversion_test.xlsx')
    
    # Run automation
    automation = L10SheetAutomation('local_conversion_test.xlsx')
    result = automation.create_next_l10_sheet_from_data(parsed_data, 'weekly')
    
    print(f"\nAutomation result: {result}")
    
    # Verify the output
    import openpyxl
    wb = openpyxl.load_workbook('local_conversion_test.xlsx')
    latest_sheet = wb[wb.sheetnames[-1]]
    
    print(f"New sheet: {latest_sheet.title}")
    print(f"Max row: {latest_sheet.max_row}")
    
    # Check AI section
    for row in range(50, min(70, latest_sheet.max_row + 1)):
        cell_value = latest_sheet.cell(row=row, column=1).value
        if cell_value and "AI IDENTIFIED" in str(cell_value):
            print(f"\nAI section found at row {row}")
            
            # Show next 10 rows
            for check_row in range(row, min(row + 10, latest_sheet.max_row + 1)):
                row_data = []
                for col in range(1, 6):
                    cell_val = latest_sheet.cell(row=check_row, column=col).value
                    if cell_val:
                        row_data.append(f"C{col}:'{str(cell_val)[:30]}{'...' if len(str(cell_val)) > 30 else ''}'")
                if row_data:
                    print(f"  Row {check_row}: {' | '.join(row_data)}")
            break
    
    wb.close()

if __name__ == "__main__":
    test_local_conversion()
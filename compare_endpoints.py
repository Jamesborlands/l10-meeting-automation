#!/usr/bin/env python3
"""
Compare the test-conversion endpoint vs process-l10 endpoint
"""

import requests
import json

your_data = {
    "new_commitments": [
        {
            "who": "Lyndsey Dunnavant",
            "task": "Ensure training policy is revised and prepared for compliance review.",
            "due_date": "Next meeting",
            "context": "To align with regulatory standards.",
            "dependencies": "Gather input from team and maintain policy visibility."
        }
    ],
    "issues_discussed": [
        {
            "issue": "The need to adjust the training policy.",
            "raised_by": "Lyndsey Dunnavant",
            "context": "The existing training policy is not being adhered to",
            "decision": "Lyndsey will redline the training policy for revision.",
            "owner": "Lyndsey Dunnavant"
        }
    ]
}

def compare_endpoints():
    """Compare both endpoints with same data"""
    print("🔍 COMPARING ENDPOINTS WITH SAME DATA")
    print("="*60)
    
    payload = {"meeting_data": your_data}
    
    # Test conversion endpoint
    print("1. Testing /test-conversion endpoint:")
    try:
        conv_response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/test-conversion',
            json=payload,
            timeout=30
        )
        
        if conv_response.status_code == 200:
            result = conv_response.json()
            print(f"   ✅ Status: {conv_response.status_code}")
            print(f"   ✅ Original keys: {result.get('original_keys')}")
            print(f"   ✅ Converted keys: {result.get('converted_keys')}")
            print(f"   ✅ NEW TO-DOS count: {result.get('new_todos_count')}")
            print(f"   ✅ Issues count: {result.get('issues_count')}")
            print(f"   ✅ Conversion successful: {result.get('conversion_successful')}")
        else:
            print(f"   ❌ Status: {conv_response.status_code}")
            print(f"   ❌ Error: {conv_response.text}")
    
    except Exception as e:
        print(f"   ❌ Error: {e}")
    
    # Test main processing endpoint
    print(f"\n2. Testing /process-l10 endpoint:")
    payload_main = {"meeting_data": your_data, "excel_url": ""}
    
    try:
        main_response = requests.post(
            'https://l10-meeting-automation-29fl.onrender.com/process-l10',
            json=payload_main,
            timeout=60
        )
        
        print(f"   ✅ Status: {main_response.status_code}")
        
        if main_response.status_code == 200:
            # Save and check Excel
            with open('endpoint_comparison.xlsx', 'wb') as f:
                f.write(main_response.content)
            
            print(f"   ✅ File size: {len(main_response.content)} bytes")
            
            # Check AI section
            import openpyxl
            wb = openpyxl.load_workbook('endpoint_comparison.xlsx')
            latest_sheet = wb[wb.sheetnames[-1]]
            
            # Count items in AI section
            items_count = 0
            for row in range(50, min(65, latest_sheet.max_row + 1)):
                for col in range(1, 6):
                    val = latest_sheet.cell(row=row, column=col).value
                    if val and str(val).strip():
                        if any(keyword in str(val) for keyword in ['Lyndsey', 'training', 'policy', 'audit']):
                            items_count += 1
                            print(f"   ✅ Found item at row {row}, col {col}: {str(val)[:50]}...")
                            break
            
            print(f"   ✅ AI section items found: {items_count}")
            wb.close()
            
        else:
            print(f"   ❌ Status: {main_response.status_code}")
            print(f"   ❌ Error: {main_response.text}")
    
    except Exception as e:
        print(f"   ❌ Error: {e}")

if __name__ == "__main__":
    compare_endpoints()